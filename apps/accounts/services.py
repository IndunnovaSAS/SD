"""
Services for accounts app - password generation, bulk upload, export, and SMS OTP.
"""

import io
import logging
import secrets
from datetime import date, timedelta

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)
User = get_user_model()


class PasswordService:
    """Service for parameterized password generation."""

    @staticmethod
    def generate_password(document_number: str, first_name: str) -> str:
        """
        Generate a standardized password from user data.

        Format: document_number + first 3 letters of first_name (lowercase)
        Example: document=1234567890, name=Carlos -> "1234567890car"
        """
        name_part = first_name.strip().lower()[:3] if first_name else "usr"
        # Pad if name is shorter than 3 chars
        name_part = name_part.ljust(3, "x")
        return f"{document_number}{name_part}"

    @staticmethod
    def reset_password(user) -> str:
        """Reset user password to the parameterized default."""
        new_password = PasswordService.generate_password(user.document_number, user.first_name)
        user.set_password(new_password)
        user.save(update_fields=["password"])
        logger.info(f"Password reset for user {user.document_number}")
        return new_password


class BulkUploadService:
    """Service for bulk user creation from Excel files."""

    REQUIRED_COLUMNS = [
        "nombre",
        "apellido",
        "numero_documento",
    ]

    OPTIONAL_COLUMNS = [
        "tipo_documento",
        "correo",
        "telefono",
        "cargo",
        "perfil_ocupacional",
        "tipo_vinculacion",
        "fecha_ingreso",
        "estado",
    ]

    DOCUMENT_TYPE_MAP = {
        "cc": "CC",
        "cédula de ciudadanía": "CC",
        "cedula de ciudadania": "CC",
        "ce": "CE",
        "cédula de extranjería": "CE",
        "ti": "TI",
        "tarjeta de identidad": "TI",
        "pasaporte": "PA",
        "pa": "PA",
    }

    JOB_PROFILE_MAP = {
        "liniero": "LINIERO",
        "técnico": "TECNICO",
        "tecnico": "TECNICO",
        "operador": "OPERADOR",
        "jefe de cuadrilla": "JEFE_CUADRILLA",
        "ingeniero residente": "INGENIERO_RESIDENTE",
        "ingeniero": "INGENIERO_RESIDENTE",
        "coordinador hseq": "COORDINADOR_HSEQ",
        "coordinador": "COORDINADOR_HSEQ",
        "administrador": "ADMINISTRADOR",
    }

    EMPLOYMENT_TYPE_MAP = {
        "directo": "direct",
        "direct": "direct",
        "contratista": "contractor",
        "contractor": "contractor",
    }

    STATUS_MAP = {
        "activo": "active",
        "active": "active",
        "inactivo": "inactive",
        "inactive": "inactive",
        "suspendido": "suspended",
        "suspended": "suspended",
        "período de prueba": "probation",
        "periodo de prueba": "probation",
        "probation": "probation",
    }

    @staticmethod
    def parse_excel(file) -> tuple[list[dict], list[str]]:
        """
        Parse an Excel file and return rows and errors.
        Returns (rows, errors) where rows is a list of dicts.
        """
        try:
            import openpyxl
        except ImportError:
            return [], ["El paquete 'openpyxl' no está instalado."]

        errors = []

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"Error al leer el archivo Excel: {e}"]

        # Read headers from first row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            val = str(cell.value or "").strip().lower()
            # Normalize header names
            val = (
                val.replace("á", "a")
                .replace("é", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ú", "u")
                .replace("ñ", "n")
                .replace(" ", "_")
            )
            headers.append(val)

        # Validate required columns
        for col in BulkUploadService.REQUIRED_COLUMNS:
            normalized = (
                col.replace("á", "a")
                .replace("é", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ú", "u")
            )
            if normalized not in headers:
                errors.append(f"Columna requerida '{col}' no encontrada en el archivo.")

        if errors:
            wb.close()
            return [], errors

        rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue  # Skip empty rows

            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value

            row_data["_row_num"] = row_num
            rows.append(row_data)

        wb.close()
        return rows, errors

    @staticmethod
    @transaction.atomic
    def create_users_from_rows(rows: list[dict]) -> tuple[list, list[str]]:
        """
        Create users from parsed rows.
        Returns (created_users, errors).
        """
        created = []
        errors = []

        for row in rows:
            row_num = row.get("_row_num", "?")

            # Required fields
            first_name = str(row.get("nombre", "")).strip()
            last_name = str(row.get("apellido", "")).strip()
            document_number = str(row.get("numero_documento", "")).strip()

            if not first_name or not last_name or not document_number:
                errors.append(f"Fila {row_num}: Nombre, apellido y documento son obligatorios.")
                continue

            # Check if user already exists
            if User.objects.filter(document_number=document_number).exists():
                errors.append(
                    f"Fila {row_num}: Ya existe un usuario con documento {document_number}."
                )
                continue

            # Optional fields
            doc_type_raw = str(row.get("tipo_documento", "CC")).strip().lower()
            document_type = BulkUploadService.DOCUMENT_TYPE_MAP.get(doc_type_raw, "CC")

            email = str(row.get("correo", "") or "").strip() or None
            phone = str(row.get("telefono", "") or "").strip()
            job_position = str(row.get("cargo", "Operario") or "Operario").strip()

            profile_raw = str(row.get("perfil_ocupacional", "LINIERO") or "LINIERO").strip().lower()
            job_profile = BulkUploadService.JOB_PROFILE_MAP.get(profile_raw, "LINIERO")

            emp_raw = str(row.get("tipo_vinculacion", "directo") or "directo").strip().lower()
            employment_type = BulkUploadService.EMPLOYMENT_TYPE_MAP.get(emp_raw, "direct")

            status_raw = str(row.get("estado", "activo") or "activo").strip().lower()
            status = BulkUploadService.STATUS_MAP.get(status_raw, "active")

            # Parse hire_date
            hire_date_raw = row.get("fecha_ingreso")
            hire_date = hire_date_raw if isinstance(hire_date_raw, date) else date.today()

            # Check email uniqueness
            if email and User.objects.filter(email=email).exists():
                errors.append(f"Fila {row_num}: Ya existe un usuario con correo {email}.")
                continue

            # Generate password
            password = PasswordService.generate_password(document_number, first_name)

            try:
                user = User(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    document_type=document_type,
                    document_number=document_number,
                    phone=phone,
                    job_position=job_position,
                    job_profile=job_profile,
                    employment_type=employment_type,
                    hire_date=hire_date,
                    status=status,
                    is_active=status == "active",
                )
                user.set_password(password)
                user.save()
                created.append(user)
            except Exception as e:
                errors.append(f"Fila {row_num}: Error al crear usuario: {e}")

        logger.info(f"Bulk upload: {len(created)} created, {len(errors)} errors")
        return created, errors

    @staticmethod
    def generate_template() -> bytes:
        """Generate a template Excel file for bulk upload."""
        try:
            import openpyxl
        except ImportError:
            return b""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        headers = [
            "nombre",
            "apellido",
            "numero_documento",
            "tipo_documento",
            "correo",
            "telefono",
            "cargo",
            "perfil_ocupacional",
            "tipo_vinculacion",
            "fecha_ingreso",
            "estado",
        ]
        ws.append(headers)

        # Example row
        ws.append(
            [
                "Juan",
                "Pérez",
                "1234567890",
                "CC",
                "juan@ejemplo.com",
                "3001234567",
                "Técnico Electricista",
                "TECNICO",
                "directo",
                date.today().isoformat(),
                "activo",
            ]
        )

        # Set column widths
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ExportService:
    """Service for exporting data to Excel."""

    @staticmethod
    def export_pending_users(category_id=None, profile=None) -> bytes:
        """
        Export users with pending/incomplete courses to Excel.
        """
        try:
            import openpyxl
        except ImportError:
            return b""

        from apps.courses.models import Enrollment

        # Get active enrollments that are NOT completed
        enrollments = (
            Enrollment.objects.filter(
                status__in=[
                    Enrollment.Status.ENROLLED,
                    Enrollment.Status.IN_PROGRESS,
                    Enrollment.Status.EXPIRED,
                ]
            )
            .select_related("user", "course", "course__category")
            .order_by("user__last_name", "user__first_name", "course__title")
        )

        # Apply filters
        if category_id:
            enrollments = enrollments.filter(course__category_id=category_id)

        if profile:
            enrollments = enrollments.filter(user__job_profile=profile)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pendientes"

        headers = [
            "Nombre",
            "Apellido",
            "Cédula",
            "Perfil Ocupacional",
            "Cargo",
            "Curso",
            "Categoría",
            "Estado",
            "Progreso (%)",
            "Fecha Inscripción",
            "Fecha Límite",
        ]
        ws.append(headers)

        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        status_display = {
            "enrolled": "Inscrito",
            "in_progress": "En progreso",
            "expired": "Vencido",
        }

        for enrollment in enrollments:
            ws.append(
                [
                    enrollment.user.first_name,
                    enrollment.user.last_name,
                    enrollment.user.document_number,
                    enrollment.user.get_job_profile_display(),
                    enrollment.user.job_position,
                    enrollment.course.title,
                    enrollment.course.category.name
                    if enrollment.course.category
                    else "Sin categoría",
                    status_display.get(enrollment.status, enrollment.status),
                    float(enrollment.progress),
                    enrollment.created_at.strftime("%Y-%m-%d") if enrollment.created_at else "",
                    enrollment.due_date.strftime("%Y-%m-%d")
                    if enrollment.due_date
                    else "Sin fecha",
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class SMSOTPService:
    """Service for SMS OTP code generation, sending, and verification."""

    OTP_LENGTH = 6
    OTP_EXPIRY_MINUTES = 5
    MAX_RESEND_PER_HOUR = 5
    MAX_VERIFY_ATTEMPTS = 5
    RESEND_COOLDOWN_SECONDS = 60

    @staticmethod
    def generate_code() -> str:
        """Generate a cryptographically secure 6-digit OTP code."""
        code = secrets.randbelow(10**SMSOTPService.OTP_LENGTH)
        return str(code).zfill(SMSOTPService.OTP_LENGTH)

    @staticmethod
    def create_otp(user, ip_address: str = None):
        """Create a new OTP code, invalidating all previous unused codes."""
        from .models import SMSOTPCode

        SMSOTPCode.objects.filter(user=user, is_used=False).update(is_used=True)

        code = SMSOTPService.generate_code()
        expires_at = timezone.now() + timedelta(minutes=SMSOTPService.OTP_EXPIRY_MINUTES)

        otp = SMSOTPCode.objects.create(
            user=user,
            code=code,
            expires_at=expires_at,
            ip_address=ip_address,
        )
        logger.info(f"OTP created for user {user.document_number} (expires at {expires_at})")
        return otp

    @staticmethod
    def send_otp(user, ip_address: str = None):
        """
        Create an OTP and send it via SMS.
        Returns (otp_instance, sent_successfully).
        Raises ValueError if rate limit exceeded or user has no phone.
        """
        if not user.phone:
            raise ValueError("El usuario no tiene número de teléfono registrado.")

        if not SMSOTPService.can_resend(user):
            raise ValueError(
                "Ha excedido el límite de envíos. Por favor, espere antes de solicitar otro código."
            )

        otp = SMSOTPService.create_otp(user, ip_address)
        sent = SMSOTPService._dispatch_sms(user, otp.code)
        return otp, sent

    @staticmethod
    def _dispatch_sms(user, code: str) -> bool:
        """Dispatch SMS: async via Celery in production, sync in dev."""
        phone = user.phone
        message = (
            f"SD LMS: Su código de verificación es {code}. "
            f"Válido por {SMSOTPService.OTP_EXPIRY_MINUTES} minutos. "
            f"No comparta este código."
        )

        if getattr(settings, "CELERY_TASK_ALWAYS_EAGER", False):
            return SMSOTPService._send_sms_sync(phone, message)
        else:
            from .tasks import send_sms_otp_task

            send_sms_otp_task.delay(phone, message)
            return True

    @staticmethod
    def _send_sms_sync(phone: str, message: str) -> bool:
        """Send SMS via Twilio. Falls back to console in dev without credentials."""
        twilio_sid = getattr(settings, "TWILIO_ACCOUNT_SID", None)
        twilio_token = getattr(settings, "TWILIO_AUTH_TOKEN", None)
        twilio_from = getattr(settings, "TWILIO_PHONE_NUMBER", None)

        if not all([twilio_sid, twilio_token, twilio_from]):
            logger.warning(f"[SMS DEV MODE] To: {phone} | Message: {message}")
            print(f"\n{'=' * 50}")
            print("SMS OTP (DEV MODE)")
            print(f"To: {phone}")
            print(f"Message: {message}")
            print(f"{'=' * 50}\n")
            return True

        try:
            from twilio.rest import Client

            client = Client(twilio_sid, twilio_token)
            client.messages.create(body=message, from_=twilio_from, to=phone)
            logger.info(f"SMS OTP sent to {phone}")
            return True
        except Exception as e:
            logger.error(f"Failed to send SMS OTP to {phone}: {e}")
            return False

    @staticmethod
    def verify_code(user, code: str) -> tuple[bool, str]:
        """
        Verify an OTP code for a user.
        Returns (is_valid, error_message).
        """
        from .models import SMSOTPCode

        try:
            otp = SMSOTPCode.objects.filter(user=user, is_used=False).latest("created_at")
        except SMSOTPCode.DoesNotExist:
            return False, "No hay código OTP pendiente. Solicite uno nuevo."

        if otp.is_expired:
            return False, "El código ha expirado. Solicite uno nuevo."

        if otp.attempts >= SMSOTPService.MAX_VERIFY_ATTEMPTS:
            otp.mark_used()
            return False, "Demasiados intentos fallidos. Solicite un nuevo código."

        if not secrets.compare_digest(otp.code, code.strip()):
            otp.increment_attempts()
            remaining = SMSOTPService.MAX_VERIFY_ATTEMPTS - otp.attempts
            return False, f"Código incorrecto. Le quedan {remaining} intento(s)."

        otp.mark_used()
        logger.info(f"OTP verified for user {user.document_number}")
        return True, ""

    @staticmethod
    def can_resend(user) -> bool:
        """Check rate limiting: max 5 OTPs per hour, 60s cooldown between sends."""
        from .models import SMSOTPCode

        now = timezone.now()
        one_hour_ago = now - timedelta(hours=1)
        recent_codes = SMSOTPCode.objects.filter(user=user, created_at__gte=one_hour_ago)

        if recent_codes.count() >= SMSOTPService.MAX_RESEND_PER_HOUR:
            return False

        last_code = recent_codes.order_by("-created_at").first()
        if last_code:
            cooldown = timedelta(seconds=SMSOTPService.RESEND_COOLDOWN_SECONDS)
            if now - last_code.created_at < cooldown:
                return False

        return True

    @staticmethod
    def get_resend_wait_seconds(user) -> int:
        """Get seconds until user can resend. Returns 0 if ready."""
        from .models import SMSOTPCode

        now = timezone.now()
        one_hour_ago = now - timedelta(hours=1)
        last_code = (
            SMSOTPCode.objects.filter(user=user, created_at__gte=one_hour_ago)
            .order_by("-created_at")
            .first()
        )

        if not last_code:
            return 0

        cooldown = timedelta(seconds=SMSOTPService.RESEND_COOLDOWN_SECONDS)
        elapsed = now - last_code.created_at
        if elapsed >= cooldown:
            return 0

        return int((cooldown - elapsed).total_seconds())

    @staticmethod
    def cleanup_expired_codes(days: int = 7) -> int:
        """Delete OTP codes older than N days."""
        from .models import SMSOTPCode

        cutoff = timezone.now() - timedelta(days=days)
        deleted, _ = SMSOTPCode.objects.filter(created_at__lt=cutoff).delete()
        if deleted:
            logger.info(f"Cleaned up {deleted} expired OTP codes")
        return deleted

    @staticmethod
    def user_requires_sms_otp(user) -> bool:
        """Determine if a user requires SMS OTP verification at login."""
        if not getattr(settings, "SMS_OTP_ENABLED", True):
            return False

        if not user.phone:
            fallback = getattr(settings, "SMS_OTP_NO_PHONE_FALLBACK", "skip")
            return fallback == "block"

        return True
